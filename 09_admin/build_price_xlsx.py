#!/usr/bin/env python3
"""RELICTUM — сборка таблицы коллекции для работы дома.

ДЕНЬГИ: только рубли, и только те суммы, что уже объявлены на сайте
(shared/catalog.js → priceValue). Валютные суммы партнёра (USD/EUR) в таблицу
НЕ попадают: по ним не подтверждено, закупочные они или продажные, а закупочные
публиковать нельзя ни при каких условиях. Всё остальное — «Цена по запросу».

Источники: shared/catalog.js (состав каталога) + 16_product_promos/promo-data.js (промо, 360°).

Запуск: python3 09_admin/build_price_xlsx.py
Выход:  09_admin/RELICTUM_коллекция_цены.xlsx
"""
import json, subprocess, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def node(expr):
    return json.loads(subprocess.run(['node', '-e', expr], cwd=ROOT,
                                     capture_output=True, text=True, check=True).stdout)

CAT = node('global.window={};require("./shared/catalog.js");console.log(JSON.stringify(window.RELICTUM_CATALOG))')
PROMO_RAW = node('global.window={};require("./16_product_promos/promo-data.js");console.log(JSON.stringify(window.RELICTUM_PROMO))')
PROMO = set(PROMO_RAW.keys())
SPINS = {k for k, v in PROMO_RAW.items() if v.get('spin')}

# Статусы по позициям: только текст, без сумм в валюте.
STATUS = {
  'R–0613': 'пинакозавр: цена — уточнить у Михаила',
  'R–0615': 'зубр: цена и валюта — уточнить; сборка под заказ',
  'R–0616': 'волк: цена и валюта — уточнить; нужен hi-res снимок',
  'R–0617': 'мамонт: нужен hi-res снимок',
  'R–0225': 'череп пещерного льва №1: цена — уточнить у Михаила',
  'R–0228': 'череп пещерного льва №2: цена и атрибуция — подтвердить у Михаила; нужен hi-res снимок и боковой ракурс',
  'R–0226': 'динокрокута: цена — ожидаем от Михаила',
  'R–0101': 'палласит Сеймчан: экспедиция в процессе, объём уточняется',
  'R–0210': 'аммолит: цена не назначена',
  'R–0103': 'Дронино индивидуал: по запросу',
  'R–0105': 'лунный метеорит в раме',
}

IVORY = 'FFF4F0E8'; BRONZE = 'FF9A6D34'; DARK = 'FF14110E'
thin = Side(style='thin', color='FFD8C8AF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook(); ws = wb.active; ws.title = 'Коллекция'
cols = ['ID', 'Экспонат', 'Латынь', 'Мир', 'Категория', 'Период', 'Регион', 'Размер',
        'Цена на сайте', 'Цена, ₽', 'Промо', '360°', 'Примечание / статус']
ws.append(cols)
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

for o in CAT:
    rub = o.get('priceValue')                      # рубли, объявленные на сайте
    ws.append([
        o['id'], o['name'], o.get('latin', ''), o.get('worldLabel', ''), o.get('category', ''),
        o.get('period', ''), o.get('region', ''), o.get('size', ''),
        o.get('price', ''), rub if isinstance(rub, (int, float)) else None,
        'да' if o['id'] in PROMO else '', 'да' if o['id'] in SPINS else '',
        STATUS.get(o['id'], ''),
    ])

for row in ws.iter_rows(min_row=2):
    for c in row:
        c.border = border; c.alignment = Alignment(vertical='top', wrap_text=True); c.font = Font(size=10)
    if row[9].value is not None:
        row[9].number_format = '#,##0 ₽'
    if row[0].row % 2 == 0:
        for c in row: c.fill = PatternFill('solid', fgColor=IVORY)

widths = [9, 32, 26, 13, 20, 14, 16, 34, 18, 16, 8, 7, 52]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

# ── лист «Сводка» ─────────────────────────────────────────────────────
s2 = wb.create_sheet('Сводка')
with_rub = [o for o in CAT if isinstance(o.get('priceValue'), (int, float))]
rows = [
    ['RELICTUM — сводка по коллекции', ''],
    ['', ''],
    ['Всего экспонатов в каталоге', len(CAT)],
    ['С промо-страницей', sum(1 for o in CAT if o['id'] in PROMO)],
    ['С 360°-видео', len(SPINS)],
    ['', ''],
    ['Цена объявлена в рублях', len(with_rub)],
    ['Сумма объявленных цен, ₽', sum(o['priceValue'] for o in with_rub)],
    ['«Цена по запросу»', len(CAT) - len(with_rub)],
    ['Ждут решения по цене', sum(1 for o in CAT if 'уточнить' in STATUS.get(o['id'], '') or 'ожидаем' in STATUS.get(o['id'], ''))],
    ['', ''],
    ['ПРИМЕЧАНИЕ', 'В таблице только рублёвые цены, объявленные на сайте. '
                   'Валютные суммы партнёра сюда не переносятся: закупочные цены не публикуются.'],
]
for r in rows: s2.append(r)
s2['A1'].font = Font(bold=True, size=14, color=BRONZE)
s2['A12'].font = Font(bold=True, color='FFB00000')
for r in s2.iter_rows(min_row=3, max_row=10, min_col=2, max_col=2):
    for c in r:
        if isinstance(c.value, (int, float)): c.number_format = '#,##0'
s2.column_dimensions['A'].width = 36; s2.column_dimensions['B'].width = 78

out = os.path.join(ROOT, '09_admin', 'RELICTUM_коллекция_цены.xlsx')
wb.save(out)
print('OK →', out, '| строк:', len(CAT), '| с рублёвой ценой:', len(with_rub))
