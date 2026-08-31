#!/usr/bin/env python3
"""RELICTUM — рабочая таблица «В наличии» для сверки документов.

Состав: только то, что показано на сайте и НЕ помечено «Под заказ»
(shared/catalog.js: status != 'Под заказ' и не hidden).

Колонки: артикул, фото, название, категория, количество, три колонки под
отметки (экспликация, квадратный сертификат, сертификат Минкульта) и ссылка.

Запуск: python3 09_admin/build_stock_xlsx.py
Выход:  09_admin/RELICTUM_наличие.xlsx
"""
import json, os, subprocess, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PImage

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMG  = os.path.join(ROOT, 'shared', 'img')
OUT  = os.path.join(ROOT, '09_admin', 'RELICTUM_наличие.xlsx')
TMP  = os.path.join(ROOT, '09_admin', '_thumbs')
SITE = 'https://relictum.gallery/objects/exhibit.html?id='

CAT = json.loads(subprocess.run(
    ['node','-e','global.window={};require("./shared/catalog.js");console.log(JSON.stringify(window.RELICTUM_CATALOG))'],
    cwd=ROOT, capture_output=True, text=True, check=True).stdout)

WORD2NUM = {'Два экземпляра':2, 'Пара половин':2, 'Шесть шаров':6, 'Десять экземпляров':10}

# Экспликация есть у всей коллекции, кроме трёх рам, добавленных 27.08.2026
# (решение Вашика 28.08). При появлении экспликаций — убирать id из набора.
NO_EXPLICATION = {'R–0622', 'R–0623', 'R–0624'}

rows = [o for o in CAT if not o.get('hidden') and o.get('status') != 'Под заказ']
rows.sort(key=lambda o: (o['category'], int(''.join(ch for ch in o['id'] if ch.isdigit()))))

INK='1A1816'; BRONZE='9A6D34'; BONE='F4F0E8'
thin = Side(style='thin', color='D8D2C6')
box  = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook(); ws = wb.active; ws.title = 'В наличии'
head = ['Артикул','Фото','Название','Категория','Кол-во, шт.',
        'Экспликация','Сертификат квадратный','Сертификат Минкульта','Ссылка на сайт']
ws.append(head)
for i,_ in enumerate(head, 1):
    c = ws.cell(row=1, column=i)
    c.font = Font(bold=True, color=BONE, size=10)
    c.fill = PatternFill('solid', fgColor=INK)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = box
ws.row_dimensions[1].height = 40

os.makedirs(TMP, exist_ok=True)
THUMB_H = 84
for n, o in enumerate(rows, start=2):
    qty = WORD2NUM.get(o.get('qty'), 1)
    ws.cell(row=n, column=1, value=o['id'])
    ws.cell(row=n, column=3, value=o['name'])
    ws.cell(row=n, column=4, value=o['category'])
    ws.cell(row=n, column=5, value=qty)
    ws.cell(row=n, column=6, value='нет' if o['id'] in NO_EXPLICATION else 'есть')
    ws.cell(row=n, column=9, value=SITE + o['slug']).hyperlink = SITE + o['slug']
    ws.cell(row=n, column=9).font = Font(color='0563C1', underline='single', size=9)
    for col in range(1, 10):
        cell = ws.cell(row=n, column=col)
        cell.border = box
        cell.alignment = Alignment(vertical='center', wrap_text=(col in (3,9)),
                                   horizontal='center' if col in (1,5,6,7,8) else 'left')
    src = os.path.join(IMG, o['img'] + '.jpg')
    ws.row_dimensions[n].height = THUMB_H * 0.78
    if os.path.exists(src):
        th = os.path.join(TMP, o['img'] + '_th.png')
        if not os.path.exists(th):
            im = PImage.open(src).convert('RGB')
            im.thumbnail((THUMB_H*2, THUMB_H*2), PImage.LANCZOS)
            im.save(th)
        pic = XLImage(th)
        k = (THUMB_H - 6) / pic.height
        pic.height = int(pic.height * k); pic.width = int(pic.width * k)
        pic.anchor = f'B{n}'
        ws.add_image(pic)

widths = {1:11, 2:16, 3:40, 4:20, 5:11, 6:15, 7:22, 8:24, 9:46}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{len(rows)+1}'

# сводка
ws2 = wb.create_sheet('Сводка')
ws2.append(['Показатель','Значение'])
for i in (1,2):
    c=ws2.cell(row=1,column=i); c.font=Font(bold=True,color=BONE,size=10); c.fill=PatternFill('solid',fgColor=INK)
total_pieces = sum(WORD2NUM.get(o.get('qty'),1) for o in rows)
ws2.append(['Позиций в наличии', len(rows)])
ws2.append(['Предметов с учётом кратных', total_pieces])
ws2.append(['Позиций с несколькими экземплярами', sum(1 for o in rows if o.get('qty'))])
ws2.append(['Всего в каталоге сайта', sum(1 for o in CAT if not o.get('hidden'))])
ws2.append(['Из них «Под заказ»', sum(1 for o in CAT if not o.get('hidden') and o.get('status')=='Под заказ')])
ws2.column_dimensions['A'].width = 42; ws2.column_dimensions['B'].width = 14

wb.save(OUT)
print('OK ·', len(rows), 'позиций ·', total_pieces, 'предметов →', OUT)
